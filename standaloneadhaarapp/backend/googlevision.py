"""OCR with PDF/TIFF as source files on GCS"""
import json
import os
import re
import traceback

from google.cloud import storage
from google.cloud import vision

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = r'backend/focus-sequencer-293417-50218e580033.json'
# gs://amul/aadhar cards (2).pdf
client = vision.ImageAnnotatorClient()
batch_size = 2
mime_type = 'application/pdf'
# gcs_source_uri='gs://amul/aadhar cards (8).pdf'

storage_client = storage.Client()
bucket = storage_client.get_bucket('harshitgoel')



def upload_file_to_gs(file, filename):
    blob = bucket.blob(filename)
    blob.upload_from_file(file)
    return 'gs://harshitgoel/' + blob.name


def async_detect_document(gcs_source_uri, gcs_destination_uri):

    feature = vision.Feature(
        type_=vision.Feature.Type.DOCUMENT_TEXT_DETECTION)

    gcs_source = vision.GcsSource(uri=gcs_source_uri)
    input_config = vision.InputConfig(
        gcs_source=gcs_source, mime_type=mime_type)

    gcs_destination = vision.GcsDestination(uri=gcs_destination_uri)
    output_config = vision.OutputConfig(batch_size=batch_size, gcs_destination=gcs_destination)

    async_request = vision.AsyncAnnotateFileRequest(
        features=[feature], input_config=input_config,
        output_config=output_config)

    operation = client.async_batch_annotate_files(
        requests=[async_request])

    print('Waiting for the operation to finish.')
    operation.result(timeout=420)

    # Once the request has completed and the output has been
    # written to GCS, we can list all the output files.
    storage_client = storage.Client()

    match = re.match(r'gs://([^/]+)/(.+)', gcs_destination_uri)
    bucket_name = match.group(1)
    prefix = match.group(2)

    bucket = storage_client.get_bucket(bucket_name)

    # List objects with the given prefix.
    blob_list = list(bucket.list_blobs(prefix=prefix))
    print('Output files:')
    for blob in blob_list:
        print(blob.name)

    # Process the first output file from GCS.
    # Since we specified batch_size=2, the first response contains
    # the first two pages of the input file.
    output = blob_list[0]

    json_string = output.download_as_string()
    json_string.decode('iso8859-1')
    response = json.loads(json_string)

    # The actual response for the first page of the input file.
    first_page_response = response['responses'][0]
    annotation = first_page_response['fullTextAnnotation']

    #NAMEHINDI
    text = {}
    text['raw'] = annotation['text']
    namehindi = None
    try:
        newlist1 = []
        for xx in annotation['text'].split('\n'):
            newlist1.append(xx)
            newlist1 = list(filter(lambda x: len(x) > 1, newlist1))
        a = 0
        str2 = "To"

        for no in newlist1:
            if str2 in no:
                b = a
            a = a + 1
        namehindi = newlist1[b + 1]
        text['namehindi'] = namehindi

    # #NAMEENGLISH
    #     translator = Translator()
    #     print(namehindi)
    #     c = translator.translate(namehindi,dest='en')
    #     wordlist = text['raw'].split("\n")
    #
    #     name = get_close_matches(c.text, wordlist)
    #     text['NameEnglish'] = name[0]


    #GENDER
        gender = []
        female_str = {"Female", "महिला", "FEMALE", "స్త్రీ"}
        male_str = {"Male", "పురుషుడు", "MALE", "ਮਰਦ", "पुरुष", "male"}
        for wordlist in text['raw'].split('\n'):
            for g in female_str:
                if re.search(g, wordlist):

                    if g not in gender:
                        print(g)
                        gender.append(g)
        if (len(gender) == 0):

            for wordlist in text['raw'].split('\n'):
                for g in male_str:
                    if re.search(g, wordlist):
                        if g not in gender:
                            print("MALE" + g)
                            gender.append(g)

        if gender[0] == "MALE" or gender[0] == "FEMALE" or gender[0]=="Male":
            gender_string = gender[1] + "/" + gender[0]
        else:
            gender_string = gender[0] + "/" + gender[1]

        text["gender string"] = gender_string

    #Download date
        match = re.search(r'Dow\w+ Date[ :]*\d+[ -/]\d+[ -/]\d+', text['raw'])
        if (match != None):
            text["Downloaddate"] = match.group()
        else:
            pass

    #Issue date
        m = re.search(r'Iss\w+ Date[ :]*\d+[ -/]\d+[ -/]\d+', text['raw'])
        if (m != None):
            text["Issuedate"] = m.group()
        else:
            pass

    #ENG ADDRESS
        addres_hin = None
        try:
            newlist = []
            for xx in text['raw'].split('\n'):
                newlist.append(xx)
                newlist = list(filter(lambda x: len(x) > 0, newlist))
                a = 0
                str = "Address:"

                for no in newlist:
                    a = a + 1
                    c = re.search(r"(?<!\d)\d{6}(?!\d)", no)
                    # r"\(\d[- \d()]\d", line)[0]

                    if c:
                        d = a
                    if str in no:
                        b = a

            addres_hin = newlist[b]
            while (b < d - 1):
                addres_hin = addres_hin + "\n" + newlist[b + 1]
                b = b + 1

        except Exception:
            pass
        text['engAddress'] = addres_hin

    #VID
        g = None
        try:
            newlist = []
            for xx in text['raw'].split('\n'):
                newlist.append(xx)
            newlist = list(filter(lambda x: len(x) > 12, newlist))
            for no in newlist:
                if re.match("^[VID : 0-9]+$", no):
                    g = no
                    g = g.replace("VID:", "")
                    g = g.replace(" ", "")

                    g = ' '.join(re.findall(r'.{1,4}', g))


        except Exception:
            pass
        text['VID'] = g
    #ADHAAR NO
        aadharno = None
        try:
            newlist = []
            str = "XXXX"
            for xx in text['raw'].split('\n'):
                newlist.append(xx)
                newlist = list(filter(lambda x: len(x) > 11, newlist))
                for word in newlist:
                    if re.match("^[0-9 ]+$", word) or str in word and len(word) == 12:
                        aadharno = word
                        aadharno = aadharno.replace(" ", "")
                        aadharno = ' '.join(re.findall(r'.{1,4}', aadharno))
        except Exception:
            pass
        text['Adhaar no'] = aadharno

    #DOB:
        wordlist=None
        birth_str = {"जन्म तिथि", "DOB", "ਜਨਮ ਮਿਤੀ", "పుట్టిన తేదీ", "DOB:", "పుట్టిన తిథి:", "Date of Birth"}
        for i in birth_str:
            for wordlist in text['raw'].split('\n'):
                if re.search(i, wordlist):
                    text["DOB"] = wordlist
                    pass

        # address hindi
        address = None
        try:
            newlist = []
            for xx in text['raw'].split('\n'):
                newlist.append(xx)
                newlist = list(filter(lambda x: len(x) > 0, newlist))
                a = 0
                str_a=""
                str = ["पता:", "ਪਤਾ:","पत्ता","पत्ता:","చిరునామా:"]
                b = 0
                d = 0
                for no in newlist:
                    a = a + 1
                    c = re.search(r"(?<!\d)\d{6}(?!\d)", no)

                    if c:
                        d = a
                    for i in str:
                        if i in no:
                            str_a = i
                            b = a
                            print(b)
                    if d > b & b != 0:
                        break
                if d > b & b != 0:
                    break

            address = newlist[b]
            while (b < d - 1):
                address = address + "\n" + newlist[b + 1]
                b = b + 1
            text['hindiAddress'] = str_a + "\n" + address
        except Exception as e:
            print(traceback.print_exc())
            pass

        phone = None
        try:

            newlist = []

            for xx in text['raw'].split('\n'):
                newlist.append(xx)
                newlist = list(filter(lambda x: len(x) > 5, newlist))
                for word in newlist:
                    if re.match("^[0-9 ]+$", word) and len(word) == 10:
                        phone = word

            text['mobile no'] = phone
        except Exception:
            pass

    except Exception as e:
        print(traceback.print_exc())
        pass

    return text

def import_to_excel(text_ex):
        import openpyxl

        workbook = openpyxl.load_workbook('my_workbook1.xlsx')
        sheet = workbook['Sheet']

        col = sheet.max_column+1
        print(col)
        # openpyxl does things based on 1 instead of 0
        # del text_ex['raw']
        row = 1
        for key, values in text_ex.items():
            # Put the key in the first column for each key in the dictionary
            sheet.cell(row=row, column=1, value=key)
            # Put the element in each adjacent column for each element in the tuple
            sheet.cell(row=row, column=col, value=values)
            row += 1

        workbook.save(filename="my_workbook1.xlsx")
        workbook.close()